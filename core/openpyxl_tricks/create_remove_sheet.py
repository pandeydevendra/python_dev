import openpyxl

'''
create_sheet to add new sheet
remove_sheet to delete existing sheet
'''
wb_load = openpyxl.load_workbook('sheets.xlsx')

wb_load.create_sheet("April")

print(wb_load.sheetnames)

sheet1 = wb_load.get_sheet_by_name("January")
wb_load.remove_sheet(sheet1)

print(wb_load.sheetnames)

wb_load.create_sheet("January", 0)
print(wb_load.sheetnames)

wb_load.create_sheet("May", 4)
print(wb_load.sheetnames)

wb_load.save('sheets2.xlsx')