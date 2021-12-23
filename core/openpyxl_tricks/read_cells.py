import openpyxl

wb_load = openpyxl.load_workbook('sample_01.xlsx')
sheet = wb_load.active

a1 = sheet['A1']
a2 = sheet['A2'].value
a3 = sheet.cell(row=3, column=1).value

print(a1.value)
print(a2)
print(a3)

"""
The example loads an existing xlsx file and reads three cells.

book = openpyxl.load_workbook('sample.xlsx')
The file is opened with the load_workbook method.

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)
We read the contents of the A1, A2, and A3 cells. In the third line, we use the cell method to get the value of A3 cell.

$ read_cells.py
56
43
None
"""
