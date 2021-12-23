# Openpyxl write to a cell
from openpyxl import Workbook

'''
here are two basic ways to write to a cell: 
using a key of a worksheet such as A1 or D3, 
         or 
using a row and column notation with the cell method.
'''
book = Workbook()
sheet = book.active

# method 1:
sheet['A3'] = 'ajinkya'
sheet['B3'] = 400

# method 2:

sheet.cell(row=4, column=1).value = 'rohit'
sheet.cell(row=5, column=1).value = 'virat'
sheet.cell(row=4, column=2).value = 150
sheet.cell(row=5, column=2).value = 220

book.save('write2cell.xlsx')

"""
from openpyxl import Workbook

book = Workbook()
sheet = book.active

sheet['A1'] = 1
sheet.cell(row=2, column=2).value = 2

book.save('write2cell.xlsx')
In the example, we write two values to two cells.

sheet['A1'] = 1
Here, we assing a numerical value to the A1 cell.

sheet.cell(row=2, column=2).value = 2
In this line, we write to cell B2 with the row and column notation.


"""
